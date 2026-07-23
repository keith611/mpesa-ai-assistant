"""
Report generation service: builds PDF and Excel report files in-memory,
ready to be streamed back via the API or attached to a WhatsApp message.
"""
import io
from datetime import datetime, timezone

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from app.db_engine import users as user_engine
from app.db_engine import transactions as txn_engine


def _fmt_money(amount) -> str:
    try:
        return f"KES {float(amount):,.2f}"
    except (TypeError, ValueError):
        return "KES 0.00"


def build_pdf_report(user_id: str, date_from: str, date_to: str) -> bytes:
    """Builds a formatted PDF statement for a user's transactions in the given period."""
    user = user_engine.get_user_by_id(user_id)
    summary = txn_engine.spending_summary(user_id, date_from, date_to)
    txns = txn_engine.search_transactions(user_id=user_id, date_from=date_from, date_to=date_to, page_size=200)["transactions"]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleStyle", parent=styles["Title"], fontSize=18, spaceAfter=4)
    subtitle_style = ParagraphStyle("SubtitleStyle", parent=styles["Normal"], fontSize=10, textColor=colors.grey)

    story = [
        Paragraph("M-Pesa AI Assistant — Financial Statement", title_style),
        Paragraph(f"{user.get('Full Name', user_id) if user else user_id} &nbsp;|&nbsp; {date_from} to {date_to}", subtitle_style),
        Spacer(1, 10 * mm),
    ]

    # Summary table
    summary_data = [
        ["Total Income", _fmt_money(summary["total_income"])],
        ["Total Expenses", _fmt_money(summary["total_spent"])],
        ["Net", _fmt_money(summary["total_income"] - summary["total_spent"])],
        ["Transactions", str(summary["transaction_count"])],
    ]
    summary_table = Table(summary_data, colWidths=[80 * mm, 80 * mm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F0F4F8")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 6 * mm))

    if summary["by_category"]:
        story.append(Paragraph("Spending by Category", styles["Heading2"]))
        cat_data = [["Category", "Amount"]] + [
            [cat, _fmt_money(amt)] for cat, amt in sorted(summary["by_category"].items(), key=lambda x: -x[1])
        ]
        cat_table = Table(cat_data, colWidths=[90 * mm, 70 * mm])
        cat_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(cat_table)
        story.append(Spacer(1, 6 * mm))

    if txns:
        story.append(Paragraph("Transaction Detail", styles["Heading2"]))
        txn_data = [["Date", "Type", "Amount", "Category", "Code"]]
        for t in txns:
            txn_data.append([
                str(t.get("Date", "")), str(t.get("Transaction Type", "")),
                _fmt_money(t.get("Amount")), str(t.get("Category", "")), str(t.get("Transaction Code", "")),
            ])
        txn_table = Table(txn_data, colWidths=[25 * mm, 30 * mm, 30 * mm, 30 * mm, 40 * mm], repeatRows=1)
        txn_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(txn_table)
    else:
        story.append(Paragraph("No transactions in this period.", styles["Normal"]))

    story.append(Spacer(1, 10 * mm))
    story.append(Paragraph(
        f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} — M-Pesa AI Assistant",
        subtitle_style,
    ))

    doc.build(story)
    return buf.getvalue()


def build_excel_report(user_id: str, date_from: str, date_to: str) -> bytes:
    """Builds a formatted (not just raw dump) .xlsx statement for a user."""
    user = user_engine.get_user_by_id(user_id)
    summary = txn_engine.spending_summary(user_id, date_from, date_to)
    txns = txn_engine.search_transactions(user_id=user_id, date_from=date_from, date_to=date_to, page_size=10000)["transactions"]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        summary_df = pd.DataFrame([
            {"Metric": "User", "Value": user.get("Full Name", user_id) if user else user_id},
            {"Metric": "Period", "Value": f"{date_from} to {date_to}"},
            {"Metric": "Total Income", "Value": summary["total_income"]},
            {"Metric": "Total Expenses", "Value": summary["total_spent"]},
            {"Metric": "Net", "Value": summary["total_income"] - summary["total_spent"]},
            {"Metric": "Transaction Count", "Value": summary["transaction_count"]},
        ])
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        txns_df = pd.DataFrame(txns) if txns else pd.DataFrame(
            columns=["Transaction ID", "Date", "Time", "Transaction Type", "Amount", "Category"])
        txns_df.to_excel(writer, sheet_name="Transactions", index=False)

        # Formatting pass
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col_idx, cell in enumerate(ws[1], start=1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(col_idx)].width = 20

    return buf.getvalue()
